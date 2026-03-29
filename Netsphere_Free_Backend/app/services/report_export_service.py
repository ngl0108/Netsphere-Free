import io
import json
from datetime import datetime
from typing import Any, Iterable


def _safe_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v)


def _pdf_escape(value: Any) -> str:
    text = _safe_text(value)
    return text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _build_minimal_pdf(lines: Iterable[Any], *, title: str = "NetSphere Report") -> bytes:
    normalized = [str(line) for line in list(lines or []) if str(line).strip()]
    if not normalized:
        normalized = ["No report data available."]

    stream_lines = ["BT", "/F1 16 Tf", "50 800 Td", f"({_pdf_escape(title)}) Tj", "/F1 10 Tf"]
    y = 780
    for line in normalized[:90]:
        safe = _pdf_escape(line)
        stream_lines.append(f"1 0 0 1 50 {y} Tm ({safe}) Tj")
        y -= 12
        if y < 60:
            break
    stream_lines.append("ET")
    stream = "\n".join(stream_lines).encode("latin-1", errors="replace")

    objects: list[bytes] = []
    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    objects.append(b"<< /Type /Pages /Count 1 /Kids [3 0 R] >>")
    objects.append(b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>")
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    objects.append(f"<< /Length {len(stream)} >>\nstream\n".encode("ascii") + stream + b"\nendstream")

    output = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(output))
        output.extend(f"{idx} 0 obj\n".encode("ascii"))
        output.extend(obj)
        output.extend(b"\nendobj\n")

    xref_start = len(output)
    output.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        output.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.extend(
        (
            f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
            f"startxref\n{xref_start}\n%%EOF\n"
        ).encode("ascii")
    )
    return bytes(output)

def _parse_details(details_raw: Any) -> dict[str, Any]:
    if isinstance(details_raw, dict):
        return details_raw
    if isinstance(details_raw, str) and details_raw.strip():
        try:
            parsed = json.loads(details_raw)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}
    return {}


def _standard_details(details_raw: Any) -> dict[str, Any]:
    details = _parse_details(details_raw)
    standards = details.get("standards") if isinstance(details, dict) else None
    if isinstance(standards, dict):
        return standards
    return details if isinstance(details, dict) else {}


def _pdf_theme():
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

    base = getSampleStyleSheet()
    primary = colors.HexColor("#0B1F3A")
    accent = colors.HexColor("#2563EB")
    success = colors.HexColor("#16A34A")
    danger = colors.HexColor("#DC2626")
    muted = colors.HexColor("#64748B")
    border = colors.HexColor("#E2E8F0")
    bg_soft = colors.HexColor("#F8FAFC")

    styles = {
        "cover_title": ParagraphStyle(
            name="nm_cover_title",
            parent=base["Title"],
            fontName="Helvetica-Bold",
            fontSize=28,
            leading=34,
            textColor=primary,
            spaceAfter=10,
        ),
        "cover_sub": ParagraphStyle(
            name="nm_cover_sub",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=12,
            leading=16,
            textColor=muted,
            spaceAfter=18,
        ),
        "title": ParagraphStyle(
            name="nm_title",
            parent=base["Title"],
            fontName="Helvetica-Bold",
            fontSize=20,
            leading=24,
            textColor=primary,
            spaceAfter=8,
        ),
        "h1": ParagraphStyle(
            name="nm_h1",
            parent=base["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=15,
            textColor=primary,
            spaceBefore=10,
            spaceAfter=6,
        ),
        "sub": ParagraphStyle(
            name="nm_sub",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=13,
            textColor=muted,
            spaceAfter=8,
        ),
        "body": ParagraphStyle(
            name="nm_body",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            textColor=primary,
            wordWrap="CJK",
        ),
        "small": ParagraphStyle(
            name="nm_small",
            parent=base["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=11,
            textColor=muted,
            wordWrap="CJK",
        ),
        "badge": ParagraphStyle(
            name="nm_badge",
            parent=base["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white,
            alignment=1,
        ),
    }

    return {
        "colors": {
            "primary": primary,
            "accent": accent,
            "success": success,
            "danger": danger,
            "muted": muted,
            "border": border,
            "bg_soft": bg_soft,
        },
        "styles": styles,
    }


def _draw_page_header_footer(canvas, doc, title: str):
    from reportlab.lib.units import mm

    theme = _pdf_theme()
    primary = theme["colors"]["primary"]
    muted = theme["colors"]["muted"]
    border = theme["colors"]["border"]

    canvas.saveState()

    w, h = doc.pagesize
    top = h - doc.topMargin + 8 * mm
    left = doc.leftMargin
    right = w - doc.rightMargin

    canvas.setLineWidth(0.5)
    canvas.setStrokeColor(border)
    canvas.line(left, top - 12 * mm, right, top - 12 * mm)

    canvas.setFillColor(primary)
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawString(left, top - 8 * mm, title)

    canvas.setFillColor(muted)
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(right, top - 8 * mm, f"Page {canvas.getPageNumber()}")

    canvas.setFillColor(muted)
    canvas.setFont("Helvetica", 7)
    canvas.drawString(left, doc.bottomMargin - 10, "Generated by NetSphere")
    canvas.drawRightString(right, doc.bottomMargin - 10, datetime.now().isoformat(timespec="seconds"))

    canvas.restoreState()


def _make_kpi_cards(rows: list[tuple[str, str, str]], total_width: float):
    from reportlab.platypus import Paragraph, Table, TableStyle
    from reportlab.lib import colors

    theme = _pdf_theme()
    border = theme["colors"]["border"]
    bg_soft = theme["colors"]["bg_soft"]
    primary = theme["colors"]["primary"]
    muted = theme["colors"]["muted"]
    body = theme["styles"]["body"]

    cells = []
    for title, value, subtitle in rows:
        cells.append(
            Paragraph(
                f"<font color='{muted.hexval()}' size='8'>{_safe_text(title)}</font><br/>"
                f"<font color='{primary.hexval()}' size='15'><b>{_safe_text(value)}</b></font><br/>"
                f"<font color='{muted.hexval()}' size='8'>{_safe_text(subtitle)}</font>",
                body,
            )
        )

    col_width = total_width / max(1, len(cells))
    table = Table([cells], colWidths=[col_width] * len(cells))
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), bg_soft),
                ("BOX", (0, 0), (-1, -1), 0.6, border),
                ("INNERGRID", (0, 0), (-1, -1), 0.6, border),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def _draw_cover_background(canvas, doc):
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    theme = _pdf_theme()
    accent = theme["colors"]["accent"]
    bg_soft = theme["colors"]["bg_soft"]
    border = theme["colors"]["border"]

    canvas.saveState()
    w, h = doc.pagesize
    canvas.setFillColor(bg_soft)
    canvas.rect(0, 0, w, h, stroke=0, fill=1)
    canvas.setFillColor(accent)
    canvas.rect(0, h - 28 * mm, w, 28 * mm, stroke=0, fill=1)
    canvas.setStrokeColor(border)
    canvas.setLineWidth(1)
    canvas.rect(doc.leftMargin - 6 * mm, doc.bottomMargin - 6 * mm, doc.width + 12 * mm, doc.height + 12 * mm, stroke=1, fill=0)
    canvas.restoreState()


def _toc_flowable():
    from reportlab.platypus import Paragraph
    from reportlab.platypus.tableofcontents import TableOfContents

    theme = _pdf_theme()
    styles = theme["styles"]

    toc = TableOfContents()
    toc.levelStyles = [
        styles["body"],
    ]
    return Paragraph("Contents", styles["title"]), toc


def _build_doc(buf: io.BytesIO, pagesize, doc_title: str, header_title: str, *, left_margin, right_margin, top_margin, bottom_margin):
    from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate

    class _Doc(BaseDocTemplate):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._header_title = header_title

        def afterFlowable(self, flowable):
            try:
                from reportlab.platypus import Paragraph
            except Exception:
                return
            if not isinstance(flowable, Paragraph):
                return
            style_name = getattr(getattr(flowable, "style", None), "name", "")
            if style_name != "nm_h1":
                return
            text = getattr(flowable, "getPlainText", lambda: "")()
            if text:
                self.notify("TOCEntry", (0, text, self.page))

    doc = _Doc(
        buf,
        pagesize=pagesize,
        title=doc_title,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="F")
    doc.addPageTemplates(
        [
            PageTemplate(id="cover", frames=[frame], onPage=lambda c, d: _draw_cover_background(c, d)),
            PageTemplate(id="body", frames=[frame], onPage=lambda c, d: _draw_page_header_footer(c, d, header_title)),
        ]
    )
    return doc


def build_inventory_xlsx(device_name: str, items: Iterable[Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    header = [
        "ent_physical_index",
        "parent_index",
        "class_id",
        "class_name",
        "name",
        "description",
        "model_name",
        "serial_number",
        "mfg_name",
        "hardware_rev",
        "firmware_rev",
        "software_rev",
        "is_fru",
        "last_seen",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i in items:
        ws.append(
            [
                getattr(i, "ent_physical_index", None),
                getattr(i, "parent_index", None),
                getattr(i, "class_id", None),
                _safe_text(getattr(i, "class_name", None)),
                _safe_text(getattr(i, "name", None)),
                _safe_text(getattr(i, "description", None)),
                _safe_text(getattr(i, "model_name", None)),
                _safe_text(getattr(i, "serial_number", None)),
                _safe_text(getattr(i, "mfg_name", None)),
                _safe_text(getattr(i, "hardware_rev", None)),
                _safe_text(getattr(i, "firmware_rev", None)),
                _safe_text(getattr(i, "software_rev", None)),
                _safe_text(getattr(i, "is_fru", None)),
                _safe_text(getattr(i, "last_seen", None)),
            ]
        )

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_compliance_xlsx(reports: Iterable[dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    summary_header = ["device_id", "device_name", "status", "score", "last_checked"]
    ws.append(summary_header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in reports:
        ws.append(
            [
                r.get("device_id"),
                _safe_text(r.get("device_name")),
                _safe_text(r.get("status")),
                r.get("score"),
                _safe_text(r.get("last_checked")),
            ]
        )

    ws2 = wb.create_sheet("Details")
    details_header = ["device_id", "standard", "total", "passed", "score", "violations"]
    ws2.append(details_header)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for r in reports:
        details = _standard_details(r.get("details"))

        if not details:
            continue

        for std_name, d in details.items():
            ws2.append(
                [
                    r.get("device_id"),
                    _safe_text(std_name),
                    d.get("total"),
                    d.get("passed"),
                    d.get("score"),
                    _safe_text(d.get("violations")),
                ]
            )

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_inventory_pdf(device_name: str, items: Iterable[Any]) -> bytes:
    from collections import Counter

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak, NextPageTemplate, KeepTogether

    buf = io.BytesIO()
    doc = _build_doc(
        buf,
        pagesize=landscape(A4),
        doc_title=f"Inventory - {device_name}",
        header_title="Inventory Report",
        left_margin=16 * mm,
        right_margin=16 * mm,
        top_margin=18 * mm,
        bottom_margin=16 * mm,
    )
    theme = _pdf_theme()
    styles = theme["styles"]
    border = theme["colors"]["border"]
    bg_soft = theme["colors"]["bg_soft"]
    primary = theme["colors"]["primary"]

    items_list = list(items or [])
    total = len(items_list)
    models = set(_safe_text(getattr(i, "model_name", "")).strip() for i in items_list if _safe_text(getattr(i, "model_name", "")).strip())
    serials = set(_safe_text(getattr(i, "serial_number", "")).strip() for i in items_list if _safe_text(getattr(i, "serial_number", "")).strip())
    class_counts = Counter(_safe_text(getattr(i, "class_name", "")).strip() or "Unknown" for i in items_list)

    elements = []
    elements.append(NextPageTemplate("body"))
    elements.append(Paragraph("Inventory Report", styles["cover_title"]))
    elements.append(Paragraph(f"Device: <b>{_safe_text(device_name)}</b>", styles["cover_sub"]))
    elements.append(
        _make_kpi_cards(
            [
                ("TOTAL ITEMS", f"{total}", "ENTITY-MIB entries"),
                ("UNIQUE MODELS", f"{len(models)}", "Detected hardware models"),
                ("SERIALS", f"{len(serials)}", "Unique serial numbers"),
            ],
            doc.width,
        )
    )
    elements.append(Spacer(1, 14))
    elements.append(PageBreak())

    toc_title, toc = _toc_flowable()
    elements.append(toc_title)
    elements.append(Spacer(1, 6))
    elements.append(toc)
    elements.append(PageBreak())

    top_classes = sorted(class_counts.items(), key=lambda x: (-x[1], x[0]))[:12]
    if top_classes:
        elements.append(Paragraph("Inventory Summary", styles["h1"]))
        summary_data = [["Class", "Count"]]
        for k, v in top_classes:
            summary_data.append([Paragraph(_safe_text(k), styles["body"]), str(v)])
        summary_table = Table(summary_data, colWidths=[doc.width * 0.7, doc.width * 0.3], repeatRows=1)
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), primary),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.4, border),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [bg_soft, colors.white]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(summary_table)
        elements.append(PageBreak())

    elements.append(Paragraph("Inventory Details", styles["h1"]))

    header = ["Index", "Parent", "Class", "Name", "Model", "Serial", "HW", "FW", "SW", "Last Seen"]
    data = [header]
    for i in items_list:
        data.append(
            [
                _safe_text(getattr(i, "ent_physical_index", "")),
                _safe_text(getattr(i, "parent_index", "")),
                Paragraph(_safe_text(getattr(i, "class_name", "")), styles["body"]),
                Paragraph(_safe_text(getattr(i, "name", "")), styles["body"]),
                Paragraph(_safe_text(getattr(i, "model_name", "")), styles["body"]),
                Paragraph(_safe_text(getattr(i, "serial_number", "")), styles["body"]),
                _safe_text(getattr(i, "hardware_rev", "")),
                _safe_text(getattr(i, "firmware_rev", "")),
                _safe_text(getattr(i, "software_rev", "")),
                _safe_text(getattr(i, "last_seen", "")),
            ]
        )

    col_widths = [
        doc.width * 0.06,
        doc.width * 0.06,
        doc.width * 0.10,
        doc.width * 0.18,
        doc.width * 0.14,
        doc.width * 0.14,
        doc.width * 0.06,
        doc.width * 0.06,
        doc.width * 0.06,
        doc.width * 0.14,
    ]
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("GRID", (0, 0), (-1, -1), 0.35, border),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("TEXTCOLOR", (0, 1), (-1, -1), primary),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    elements.append(KeepTogether([table]))
    doc.multiBuild(elements)
    return buf.getvalue()


def build_compliance_pdf(reports: Iterable[dict[str, Any]]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak, NextPageTemplate, KeepTogether

    buf = io.BytesIO()
    doc = _build_doc(
        buf,
        pagesize=A4,
        doc_title="Compliance Reports",
        header_title="Compliance Report",
        left_margin=16 * mm,
        right_margin=16 * mm,
        top_margin=18 * mm,
        bottom_margin=16 * mm,
    )
    theme = _pdf_theme()
    styles = theme["styles"]
    primary = theme["colors"]["primary"]
    border = theme["colors"]["border"]
    bg_soft = theme["colors"]["bg_soft"]
    success = theme["colors"]["success"]
    danger = theme["colors"]["danger"]
    muted = theme["colors"]["muted"]

    reports_list = list(reports or [])
    total = len(reports_list)
    compliant = sum(1 for r in reports_list if str(r.get("status") or "").lower() == "compliant")
    violations = sum(1 for r in reports_list if str(r.get("status") or "").lower() == "violation")
    avg_score = 0.0
    if total:
        scores = []
        for r in reports_list:
            try:
                scores.append(float(r.get("score")))
            except Exception:
                pass
        avg_score = sum(scores) / len(scores) if scores else 0.0

    elements = []
    elements.append(NextPageTemplate("body"))
    elements.append(Paragraph("Compliance Report", styles["cover_title"]))
    elements.append(Paragraph("Rule-based compliance scan results and drift indicators.", styles["cover_sub"]))
    elements.append(
        _make_kpi_cards(
            [
                ("DEVICES", f"{total}", "Reports available"),
                ("COMPLIANT", f"{compliant}", "No violations detected"),
                ("VIOLATIONS", f"{violations}", "Action required"),
                ("AVG SCORE", f"{avg_score:.1f}%", "Overall posture"),
            ],
            doc.width,
        )
    )
    elements.append(Spacer(1, 14))
    elements.append(PageBreak())

    toc_title, toc = _toc_flowable()
    elements.append(toc_title)
    elements.append(Spacer(1, 6))
    elements.append(toc)
    elements.append(PageBreak())

    elements.append(Paragraph("Summary", styles["h1"]))

    data = [["Device", "Status", "Score", "Last Checked"]]
    for r in sorted(reports_list, key=lambda x: (_safe_text(x.get("device_name") or x.get("device_id")).lower())):
        status = _safe_text(r.get("status")).lower()
        status_label = status.upper() if status else "UNKNOWN"
        data.append(
            [
                Paragraph(_safe_text(r.get("device_name") or r.get("device_id")), styles["body"]),
                Paragraph(status_label, styles["body"]),
                _safe_text(r.get("score")),
                _safe_text(r.get("last_checked")),
            ]
        )

    table = Table(data, repeatRows=1, colWidths=[doc.width * 0.40, doc.width * 0.18, doc.width * 0.14, doc.width * 0.28])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, border),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    for idx in range(1, len(data)):
        status = str(getattr(data[idx][1], "getPlainText", lambda: data[idx][1])() or "").lower()
        if status == "compliant":
            table.setStyle(
                TableStyle(
                    [
                        ("TEXTCOLOR", (1, idx), (1, idx), success),
                        ("FONTNAME", (1, idx), (1, idx), "Helvetica-Bold"),
                        ("BACKGROUND", (1, idx), (1, idx), colors.HexColor("#ECFDF5")),
                    ]
                )
            )
        elif status == "violation":
            table.setStyle(
                TableStyle(
                    [
                        ("TEXTCOLOR", (1, idx), (1, idx), danger),
                        ("FONTNAME", (1, idx), (1, idx), "Helvetica-Bold"),
                        ("BACKGROUND", (1, idx), (1, idx), colors.HexColor("#FEF2F2")),
                    ]
                )
            )

    elements.append(table)

    def _extract_violation_rows(details: dict[str, Any]) -> list[list[str]]:
        out = []
        for std_name, d in _standard_details(details).items():
            violations_list = d.get("violations") if isinstance(d, dict) else None
            if not isinstance(violations_list, list):
                continue
            for v in violations_list:
                if not isinstance(v, dict):
                    continue
                out.append(
                    [
                        _safe_text(std_name),
                        _safe_text(v.get("rule")),
                        _safe_text(v.get("severity")),
                        _safe_text(v.get("remediation")),
                    ]
                )
        return out

    violators = [r for r in reports_list if str(r.get("status") or "").lower() == "violation"]
    if violators:
        elements.append(PageBreak())
        elements.append(Paragraph("Findings (Top Violations)", styles["h1"]))
        for r in sorted(violators, key=lambda x: (_safe_text(x.get("device_name") or x.get("device_id")).lower())):
            dev_name = _safe_text(r.get("device_name") or r.get("device_id"))
            elements.append(Paragraph(f"<b>{dev_name}</b>", styles["body"]))
            details = _standard_details(r.get("details"))
            rows = _extract_violation_rows(details)[:20]
            if not rows:
                elements.append(Paragraph("No detailed violations available.", styles["small"]))
                elements.append(Spacer(1, 6))
                continue
            vdata = [["Standard", "Rule", "Severity", "Remediation"]] + rows
            vtable = Table(vdata, repeatRows=1, colWidths=[doc.width * 0.18, doc.width * 0.32, doc.width * 0.12, doc.width * 0.38])
            vtable.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), primary),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 0.35, border),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            for ridx in range(1, len(vdata)):
                sev = str(vdata[ridx][2] or "").strip().lower()
                if sev in {"critical", "high"}:
                    vtable.setStyle(TableStyle([("BACKGROUND", (2, ridx), (2, ridx), colors.HexColor("#FEE2E2")), ("TEXTCOLOR", (2, ridx), (2, ridx), danger), ("FONTNAME", (2, ridx), (2, ridx), "Helvetica-Bold")]))
                elif sev in {"warning", "medium"}:
                    vtable.setStyle(TableStyle([("BACKGROUND", (2, ridx), (2, ridx), colors.HexColor("#FEF9C3")), ("TEXTCOLOR", (2, ridx), (2, ridx), colors.HexColor("#A16207")), ("FONTNAME", (2, ridx), (2, ridx), "Helvetica-Bold")]))
                elif sev in {"low", "info"}:
                    vtable.setStyle(TableStyle([("BACKGROUND", (2, ridx), (2, ridx), colors.HexColor("#E0F2FE")), ("TEXTCOLOR", (2, ridx), (2, ridx), colors.HexColor("#0369A1")), ("FONTNAME", (2, ridx), (2, ridx), "Helvetica-Bold")]))
            elements.append(KeepTogether([vtable]))
            elements.append(Spacer(1, 10))

    doc.multiBuild(elements)
    return buf.getvalue()


def build_preventive_check_pdf(run_payload: dict[str, Any]) -> bytes:
    payload = dict(run_payload or {})
    summary = dict(payload.get("summary") or {})
    findings = list(payload.get("findings") or [])

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import KeepTogether, NextPageTemplate, PageBreak, Paragraph, Spacer, Table, TableStyle
    except ModuleNotFoundError:
        fallback_lines = [
            f"Template: {_safe_text(payload.get('template_name') or 'Unnamed Template')}",
            f"Triggered by: {_safe_text(payload.get('triggered_by') or 'operator')}",
            f"Started at: {_safe_text(payload.get('started_at') or '-')}",
            f"Finished at: {_safe_text(payload.get('finished_at') or '-')}",
            f"Devices reviewed: {int(summary.get('devices_total') or 0)}",
            f"Critical devices: {int(summary.get('critical_devices') or 0)}",
            f"Warning devices: {int(summary.get('warning_devices') or 0)}",
            f"Failed checks: {int(summary.get('failed_checks_total') or 0)}",
            "",
            "Findings:",
        ]
        for row in findings:
            fallback_lines.append(
                f"- {_safe_text(row.get('device_name') or 'Unknown Device')} "
                f"[{_safe_text(row.get('status') or 'healthy').upper()}]"
            )
            for item in list(row.get("findings") or [])[:4]:
                fallback_lines.append(
                    f"  * {_safe_text(item.get('severity') or 'warning').upper()} "
                    f"{_safe_text(item.get('check_key') or 'check')}: {_safe_text(item.get('message') or '')}"
                )
        return _build_minimal_pdf(
            fallback_lines,
            title=f"Preventive Check Report - {_safe_text(payload.get('template_name') or 'Run')}",
        )

    buf = io.BytesIO()
    doc = _build_doc(
        buf,
        pagesize=A4,
        doc_title=f"Preventive Check - {_safe_text(payload.get('template_name') or 'Run')}",
        header_title="Preventive Check Report",
        left_margin=16 * mm,
        right_margin=16 * mm,
        top_margin=18 * mm,
        bottom_margin=16 * mm,
    )
    theme = _pdf_theme()
    styles = theme["styles"]
    primary = theme["colors"]["primary"]
    border = theme["colors"]["border"]
    bg_soft = theme["colors"]["bg_soft"]
    success = theme["colors"]["success"]
    danger = theme["colors"]["danger"]
    muted = theme["colors"]["muted"]

    scope = dict(summary.get("target_scope") or {})
    management_states = ", ".join(list(scope.get("management_states") or [])) or "managed"
    roles = ", ".join(list(scope.get("roles") or [])) or "all roles"

    elements = []
    elements.append(NextPageTemplate("body"))
    elements.append(Paragraph("Preventive Check Report", styles["cover_title"]))
    elements.append(
        Paragraph(
            (
                f"Template: <b>{_safe_text(payload.get('template_name') or 'Unnamed Template')}</b><br/>"
                f"Triggered by: {_safe_text(payload.get('triggered_by') or 'operator')}<br/>"
                f"Execution window: {_safe_text(payload.get('started_at') or '-')} ~ {_safe_text(payload.get('finished_at') or '-')}"
            ),
            styles["cover_sub"],
        )
    )
    elements.append(
        _make_kpi_cards(
            [
                ("DEVICES REVIEWED", f"{int(summary.get('devices_total') or 0)}", f"Scope: {management_states}"),
                ("CRITICAL", f"{int(summary.get('critical_devices') or 0)}", "Immediate follow-up required"),
                ("WARNING", f"{int(summary.get('warning_devices') or 0)}", "Needs scheduled remediation"),
                ("FAILED CHECKS", f"{int(summary.get('failed_checks_total') or 0)}", f"Roles: {roles}"),
            ],
            doc.width,
        )
    )
    elements.append(Spacer(1, 14))
    elements.append(PageBreak())

    toc_title, toc = _toc_flowable()
    elements.append(toc_title)
    elements.append(Spacer(1, 6))
    elements.append(toc)
    elements.append(PageBreak())

    elements.append(Paragraph("Execution Summary", styles["h1"]))
    summary_data = [
        ["Field", "Value"],
        ["Run ID", _safe_text(payload.get("id"))],
        ["Template ID", _safe_text(payload.get("template_id"))],
        ["Status", _safe_text(payload.get("status") or "unknown").upper()],
        ["Triggered by", _safe_text(payload.get("triggered_by") or "operator")],
        ["Started at", _safe_text(payload.get("started_at") or "-")],
        ["Finished at", _safe_text(payload.get("finished_at") or "-")],
        ["Management scope", management_states],
        ["Role scope", roles],
        ["Healthy devices", _safe_text(summary.get("healthy_devices") or 0)],
        ["Info devices", _safe_text(summary.get("info_devices") or 0)],
        ["Warning devices", _safe_text(summary.get("warning_devices") or 0)],
        ["Critical devices", _safe_text(summary.get("critical_devices") or 0)],
        ["Failed checks total", _safe_text(summary.get("failed_checks_total") or 0)],
    ]
    summary_table = Table(summary_data, repeatRows=1, colWidths=[doc.width * 0.34, doc.width * 0.66])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, border),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(summary_table)

    elements.append(PageBreak())
    elements.append(Paragraph("Device Review", styles["h1"]))

    if not findings:
        elements.append(Paragraph("No device findings were captured in this preventive check run.", styles["body"]))
    else:
        rows = [["Device", "Preventive Status", "Role", "Management", "Findings"]]
        for row in findings:
            row_findings = list(row.get("findings") or [])
            finding_text = "No findings"
            if row_findings:
                finding_text = "<br/>".join(
                    (
                        f"<b>[{_safe_text(item.get('severity') or 'warning').upper()}]</b> "
                        f"{_safe_text(item.get('check_key') or 'check').replace('_', ' ')}: "
                        f"{_safe_text(item.get('message') or '')}"
                    )
                    for item in row_findings
                )
            rows.append(
                [
                    Paragraph(
                        f"<b>{_safe_text(row.get('device_name') or 'Unknown Device')}</b><br/>"
                        f"<font color='{muted.hexval()}'>{_safe_text(row.get('ip_address') or '-')}</font>",
                        styles["body"],
                    ),
                    Paragraph(_safe_text(row.get("status") or "healthy").upper(), styles["body"]),
                    Paragraph(_safe_text(row.get("role") or "-"), styles["body"]),
                    Paragraph(_safe_text(row.get("management_state") or "-"), styles["body"]),
                    Paragraph(finding_text, styles["body"]),
                ]
            )

        review_table = Table(
            rows,
            repeatRows=1,
            colWidths=[
                doc.width * 0.20,
                doc.width * 0.14,
                doc.width * 0.12,
                doc.width * 0.14,
                doc.width * 0.40,
            ],
        )
        review_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), primary),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.35, border),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        for idx, row in enumerate(findings, start=1):
            status = str(row.get("status") or "").strip().lower()
            if status == "critical":
                review_table.setStyle(
                    TableStyle(
                        [
                            ("TEXTCOLOR", (1, idx), (1, idx), danger),
                            ("FONTNAME", (1, idx), (1, idx), "Helvetica-Bold"),
                            ("BACKGROUND", (1, idx), (1, idx), colors.HexColor("#FEF2F2")),
                        ]
                    )
                )
            elif status == "warning":
                review_table.setStyle(
                    TableStyle(
                        [
                            ("TEXTCOLOR", (1, idx), (1, idx), colors.HexColor("#A16207")),
                            ("FONTNAME", (1, idx), (1, idx), "Helvetica-Bold"),
                            ("BACKGROUND", (1, idx), (1, idx), colors.HexColor("#FEF9C3")),
                        ]
                    )
                )
            elif status == "info":
                review_table.setStyle(
                    TableStyle(
                        [
                            ("TEXTCOLOR", (1, idx), (1, idx), colors.HexColor("#0369A1")),
                            ("FONTNAME", (1, idx), (1, idx), "Helvetica-Bold"),
                            ("BACKGROUND", (1, idx), (1, idx), colors.HexColor("#E0F2FE")),
                        ]
                    )
                )
            else:
                review_table.setStyle(
                    TableStyle(
                        [
                            ("TEXTCOLOR", (1, idx), (1, idx), success),
                            ("FONTNAME", (1, idx), (1, idx), "Helvetica-Bold"),
                            ("BACKGROUND", (1, idx), (1, idx), colors.HexColor("#ECFDF5")),
                        ]
                    )
                )
        elements.append(KeepTogether([review_table]))

    doc.multiBuild(elements)
    return buf.getvalue()


def build_operations_review_markdown(snapshot_payload: dict[str, Any]) -> str:
    payload = dict(snapshot_payload or {})
    preventive = dict(payload.get("preventive_checks") or {})
    preventive_summary = dict(preventive.get("summary") or {})
    preventive_runs = list(preventive.get("recent_runs") or [])
    approvals = dict(payload.get("approvals") or {})
    approval_summary = dict(approvals.get("summary") or {})
    approval_items = list(approvals.get("items") or [])
    service_groups = dict(payload.get("service_groups") or {})
    group_items = list(service_groups.get("items") or [])
    service_issues = dict(payload.get("service_issues") or {})
    issue_items = list(service_issues.get("items") or [])
    action_continuity = dict(payload.get("action_continuity") or {})
    action_summary = dict(action_continuity.get("summary") or {})
    action_items = list(action_continuity.get("items") or [])
    follow_up_agenda = dict(payload.get("follow_up_agenda") or {})
    follow_up_summary = dict(follow_up_agenda.get("summary") or {})
    follow_up_items = list(follow_up_agenda.get("items") or [])
    state_history = dict(payload.get("state_history") or {})
    state_latest = dict(state_history.get("latest_snapshot") or {})
    state_compare = dict(state_history.get("latest_compare") or {})
    state_hotspots = list(state_history.get("review_hotspots") or [])
    release = dict(payload.get("release_evidence") or {})
    release_summary = dict(release.get("summary") or {})
    release_sections = list(release.get("sections") or [])

    lines: list[str] = []
    lines.append("# NetSphere Operations Review")
    lines.append("")
    lines.append(f"- Generated at: {_safe_text(payload.get('generated_at') or '-')}")
    lines.append(f"- Preventive templates: {int(preventive_summary.get('templates_total') or 0)}")
    lines.append(f"- Recent preventive runs: {int(preventive_summary.get('recent_runs_total') or 0)}")
    lines.append(f"- Recent approvals: {int(approval_summary.get('total') or 0)}")
    lines.append(f"- Pending approvals: {int(approval_summary.get('pending') or 0)}")
    lines.append(f"- Service groups: {int(service_groups.get('total') or 0)}")
    lines.append(f"- Service-scoped issues: {int(service_issues.get('total') or 0)}")
    lines.append(f"- Issues with active action context: {int(action_summary.get('with_active_actions') or 0)}")
    lines.append(f"- Follow-up items ready for handoff: {int(follow_up_summary.get('ready_for_handoff') or 0)}")
    lines.append(f"- State snapshots: {int(state_history.get('snapshot_count') or 0)}")
    lines.append(f"- Latest state review: {_safe_text(state_compare.get('result') or 'unavailable')}")
    lines.append(f"- Release status: {_safe_text(release_summary.get('overall_status') or 'unknown')}")
    lines.append("")

    lines.append("## State History Review")
    lines.append(f"- Stored snapshots: {int(state_history.get('snapshot_count') or 0)}")
    lines.append(f"- Latest snapshot: {_safe_text(state_latest.get('label') or state_latest.get('generated_at') or '-')}")
    lines.append(f"- Latest age (hours): {_safe_text(state_latest.get('age_hours') if state_latest.get('age_hours') is not None else '-')}")
    lines.append(f"- Review result: {_safe_text(state_compare.get('result') or 'unavailable')}")
    lines.append(f"- Review cards: {int(state_compare.get('review_cards') or 0)}")
    lines.append(f"- Improved cards: {int(state_compare.get('improved_cards') or 0)}")
    lines.append(f"- Changed cards: {int(state_compare.get('changed_cards') or 0)}")
    if state_hotspots:
        lines.append("")
        lines.append("### Review Hotspots")
        for item in state_hotspots:
            lines.append(
                f"- {item.get('title') or item.get('key') or 'signal'} | "
                f"status={item.get('status') or 'steady'} | "
                f"delta={item.get('delta') or '-'} | "
                f"recommendation={item.get('recommendation') or '-'}"
            )
    lines.append("")

    lines.append("## Preventive Check Summary")
    lines.append(f"- Enabled templates: {int(preventive_summary.get('enabled_templates') or 0)} / {int(preventive_summary.get('templates_total') or 0)}")
    lines.append(f"- Recent failed checks: {int(preventive_summary.get('recent_failed_checks_total') or 0)}")
    lines.append(f"- Recent critical devices: {int(preventive_summary.get('recent_critical_devices') or 0)}")
    lines.append(f"- Last run at: {_safe_text(preventive_summary.get('last_run_at') or '-')}")
    lines.append("")
    if preventive_runs:
        lines.append("### Recent Runs")
        for run in preventive_runs:
            summary = dict(run.get("summary") or {})
            lines.append(
                f"- {run.get('template_name') or 'Unnamed Template'} | "
                f"status={run.get('status') or 'completed'} | "
                f"devices={int(summary.get('devices_total') or 0)} | "
                f"failed_checks={int(summary.get('failed_checks_total') or 0)} | "
                f"critical={int(summary.get('critical_devices') or 0)}"
            )
        lines.append("")

    lines.append("## Change Control Review")
    lines.append(f"- Pending approvals: {int(approval_summary.get('pending') or 0)}")
    lines.append(f"- Evidence ready: {int(approval_summary.get('evidence_ready_count') or 0)}")
    lines.append(f"- Rollback tracked: {int(approval_summary.get('rollback_tracked_count') or 0)}")
    lines.append("")
    if approval_items:
        lines.append("### Recent Approval Requests")
        for item in approval_items:
            status = item.get("execution_status") or item.get("status") or "pending"
            lines.append(
                f"- #{item.get('id')} {item.get('title') or 'Request'} | "
                f"type={item.get('request_type') or 'approval'} | "
                f"status={status} | requester={item.get('requester_name') or 'operator'}"
            )
        lines.append("")

    lines.append("## Service Group Snapshot")
    lines.append(f"- Groups included in this review: {len(group_items)}")
    lines.append("")
    if group_items:
        for group in group_items:
            lines.append(
                f"- {group.get('name') or 'Unnamed Group'} | "
                f"criticality={group.get('criticality') or 'standard'} | "
                f"owner={group.get('owner_team') or 'unassigned'} | "
                f"members={int(group.get('member_count') or 0)}"
            )
        lines.append("")

    lines.append("## Service-Scoped Operational Issues")
    if not issue_items:
        lines.append("- No active service-scoped issues were included in this review.")
    else:
        for issue in issue_items:
            service_summary = dict(issue.get("service_impact_summary") or {})
            action_summary = dict(issue.get("action_summary") or {})
            knowledge_summary = dict(issue.get("knowledge_summary") or {})
            sop_summary = dict(issue.get("sop_summary") or {})
            approval_ctx = dict(issue.get("approval_summary") or {})
            lines.append(
                f"- {issue.get('title') or 'Issue'} | severity={issue.get('severity') or 'info'} | "
                f"service={service_summary.get('primary_name') or '-'} | "
                f"sop={sop_summary.get('readiness_status') or 'limited_context'} | "
                f"actions={int(action_summary.get('total') or 0)} | "
                f"knowledge={int(knowledge_summary.get('recommendation_count') or 0)} | "
                f"approvals={int(approval_ctx.get('total') or 0)}"
            )
    lines.append("")

    lines.append("## Action Continuity Review")
    lines.append(f"- Issues in scope: {int(action_summary.get('issues_in_scope') or len(action_items) or 0)}")
    lines.append(f"- With active actions: {int(action_summary.get('with_active_actions') or 0)}")
    lines.append(f"- With assigned owners: {int(action_summary.get('with_assignee') or 0)}")
    lines.append(f"- With knowledge matches: {int(action_summary.get('with_knowledge') or 0)}")
    lines.append(f"- With evidence-linked approvals: {int(action_summary.get('with_evidence_ready') or 0)}")
    lines.append(f"- Limited context: {int(action_summary.get('limited_context') or 0)}")
    if action_items:
        lines.append("")
        lines.append("### Active Continuity Items")
        for item in action_items:
            lines.append(
                f"- {item.get('title') or 'Issue'} | "
                f"service={item.get('primary_service') or '-'} | "
                f"action={item.get('action_status') or 'open'} | "
                f"owner={item.get('action_owner') or '-'} | "
                f"knowledge={int(item.get('knowledge_matches') or 0)} | "
                f"sop={item.get('sop_status') or 'limited_context'} | "
                f"approvals={int(item.get('approval_total') or 0)}"
            )
    lines.append("")

    lines.append("## Follow-up Agenda")
    lines.append(f"- Needs action creation: {int(follow_up_summary.get('needs_action') or 0)}")
    lines.append(f"- Needs owner assignment: {int(follow_up_summary.get('needs_owner') or 0)}")
    lines.append(f"- Needs knowledge capture: {int(follow_up_summary.get('needs_knowledge') or 0)}")
    lines.append(f"- Needs evidence capture: {int(follow_up_summary.get('needs_evidence') or 0)}")
    lines.append(f"- Ready for handoff: {int(follow_up_summary.get('ready_for_handoff') or 0)}")
    if follow_up_items:
        lines.append("")
        lines.append("### Follow-up Items")
        for item in follow_up_items:
            lines.append(
                f"- {item.get('title') or 'Issue'} | "
                f"priority={item.get('priority') or 'normal'} | "
                f"next={item.get('recommended_step') or 'review_and_handoff'} | "
                f"service={item.get('primary_service') or '-'} | "
                f"owner={item.get('action_owner') or '-'} | "
                f"note={item.get('action_note') or '-'}"
            )
    lines.append("")

    lines.append("## Release Readiness")
    lines.append(
        f"- Accepted gates: {int(release_summary.get('accepted_gates') or 0)} / "
        f"{int(release_summary.get('available_gates') or release_summary.get('total_gates') or 0)}"
    )
    warnings = list(release_summary.get("warning_gates") or [])
    in_progress = list(release_summary.get("in_progress_gates") or [])
    if warnings:
        lines.append(f"- Warning gates: {', '.join(str(item) for item in warnings)}")
    if in_progress:
        lines.append(f"- In progress gates: {', '.join(str(item) for item in in_progress)}")
    if release_sections:
        lines.append("")
        lines.append("### Release Sections")
        for section in release_sections:
            lines.append(
                f"- {section.get('title') or section.get('id') or 'section'} | "
                f"status={section.get('status') or 'unknown'} | "
                f"summary={section.get('summary') or '-'}"
            )

    return "\n".join(lines).strip() + "\n"


def build_operations_review_pdf(snapshot_payload: dict[str, Any]) -> bytes:
    payload = dict(snapshot_payload or {})
    preventive = dict(payload.get("preventive_checks") or {})
    preventive_summary = dict(preventive.get("summary") or {})
    preventive_runs = list(preventive.get("recent_runs") or [])
    approvals = dict(payload.get("approvals") or {})
    approval_summary = dict(approvals.get("summary") or {})
    approval_items = list(approvals.get("items") or [])
    service_groups = dict(payload.get("service_groups") or {})
    group_items = list(service_groups.get("items") or [])
    service_issues = dict(payload.get("service_issues") or {})
    issue_items = list(service_issues.get("items") or [])
    action_continuity = dict(payload.get("action_continuity") or {})
    action_summary = dict(action_continuity.get("summary") or {})
    action_items = list(action_continuity.get("items") or [])
    follow_up_agenda = dict(payload.get("follow_up_agenda") or {})
    follow_up_summary = dict(follow_up_agenda.get("summary") or {})
    follow_up_items = list(follow_up_agenda.get("items") or [])
    state_history = dict(payload.get("state_history") or {})
    state_latest = dict(state_history.get("latest_snapshot") or {})
    state_compare = dict(state_history.get("latest_compare") or {})
    state_hotspots = list(state_history.get("review_hotspots") or [])
    release = dict(payload.get("release_evidence") or {})
    release_summary = dict(release.get("summary") or {})
    release_sections = list(release.get("sections") or [])

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import NextPageTemplate, PageBreak, Paragraph, Spacer, Table, TableStyle
    except ModuleNotFoundError:
        fallback_lines = build_operations_review_markdown(payload).splitlines()
        return _build_minimal_pdf(fallback_lines, title="NetSphere Operations Review")

    buf = io.BytesIO()
    doc = _build_doc(
        buf,
        pagesize=A4,
        doc_title="NetSphere Operations Review",
        header_title="Operations Review",
        left_margin=16 * mm,
        right_margin=16 * mm,
        top_margin=18 * mm,
        bottom_margin=16 * mm,
    )
    theme = _pdf_theme()
    styles = theme["styles"]
    primary = theme["colors"]["primary"]
    border = theme["colors"]["border"]
    bg_soft = theme["colors"]["bg_soft"]
    muted = theme["colors"]["muted"]

    elements = []
    elements.append(NextPageTemplate("body"))
    elements.append(Paragraph("NetSphere Operations Review", styles["cover_title"]))
    elements.append(
        Paragraph(
            (
                f"Generated at: {_safe_text(payload.get('generated_at') or '-')}<br/>"
                f"Preventive runs in scope: {int(preventive_summary.get('recent_runs_total') or 0)}<br/>"
                f"Recent approvals in scope: {int(approval_summary.get('total') or 0)}<br/>"
                f"Service-scoped issues in scope: {int(service_issues.get('total') or 0)}<br/>"
                f"Issues with active action context: {int(action_summary.get('with_active_actions') or 0)}<br/>"
                f"Follow-up items ready for handoff: {int(follow_up_summary.get('ready_for_handoff') or 0)}<br/>"
                f"State snapshots in scope: {int(state_history.get('snapshot_count') or 0)}"
            ),
            styles["cover_sub"],
        )
    )
    elements.append(
        _make_kpi_cards(
            [
                ("PREVENTIVE RUNS", f"{int(preventive_summary.get('recent_runs_total') or 0)}", f"{int(preventive_summary.get('recent_failed_checks_total') or 0)} failed checks"),
                ("PENDING APPROVALS", f"{int(approval_summary.get('pending') or 0)}", f"{int(approval_summary.get('evidence_ready_count') or 0)} evidence ready"),
                ("SERVICE ISSUES", f"{int(service_issues.get('total') or 0)}", f"{len(group_items)} groups in review"),
                ("ACTION CONTEXT", f"{int(action_summary.get('with_active_actions') or 0)}", f"{int(action_summary.get('with_assignee') or 0)} owned"),
                ("FOLLOW-UP", f"{int(follow_up_summary.get('ready_for_handoff') or 0)}", f"{int(follow_up_summary.get('needs_action') or 0)} need action"),
                ("STATE REVIEW", f"{int(state_history.get('snapshot_count') or 0)}", str(state_compare.get('result') or 'unavailable').upper()),
            ],
            doc.width,
        )
    )
    elements.append(Spacer(1, 14))
    elements.append(PageBreak())

    toc_title, toc = _toc_flowable()
    elements.append(toc_title)
    elements.append(Spacer(1, 6))
    elements.append(toc)
    elements.append(PageBreak())

    elements.append(Paragraph("State History Review", styles["h1"]))
    state_rows = [
        ["Metric", "Value"],
        ["Stored Snapshots", _safe_text(state_history.get("snapshot_count") or 0)],
        ["Latest Snapshot", _safe_text(state_latest.get("label") or state_latest.get("generated_at") or "-")],
        ["Latest Age (hours)", _safe_text(state_latest.get("age_hours") if state_latest.get("age_hours") is not None else "-")],
        ["Review Result", _safe_text(state_compare.get("result") or "unavailable").upper()],
        ["Review Cards", _safe_text(state_compare.get("review_cards") or 0)],
        ["Improved Cards", _safe_text(state_compare.get("improved_cards") or 0)],
        ["Changed Cards", _safe_text(state_compare.get("changed_cards") or 0)],
    ]
    state_table = Table(state_rows, repeatRows=1, colWidths=[doc.width * 0.42, doc.width * 0.58])
    state_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, border),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(state_table)
    if state_hotspots:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Review Hotspots", styles["h1"]))
        hotspot_rows = [["Hotspot", "Delta", "Recommendation"]]
        for item in state_hotspots[:3]:
            hotspot_rows.append(
                [
                    _safe_text(item.get("title") or item.get("key") or "signal"),
                    _safe_text(item.get("delta") or "-"),
                    _safe_text(item.get("recommendation") or "-"),
                ]
            )
        hotspot_table = Table(
            hotspot_rows,
            repeatRows=1,
            colWidths=[doc.width * 0.24, doc.width * 0.18, doc.width * 0.58],
        )
        hotspot_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), primary),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.35, border),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(hotspot_table)

    elements.append(PageBreak())
    elements.append(Paragraph("Preventive Check Summary", styles["h1"]))
    preventive_rows = [
        ["Metric", "Value"],
        ["Templates", _safe_text(preventive_summary.get("templates_total") or 0)],
        ["Enabled Templates", _safe_text(preventive_summary.get("enabled_templates") or 0)],
        ["Recent Runs", _safe_text(preventive_summary.get("recent_runs_total") or 0)],
        ["Recent Failed Checks", _safe_text(preventive_summary.get("recent_failed_checks_total") or 0)],
        ["Recent Critical Devices", _safe_text(preventive_summary.get("recent_critical_devices") or 0)],
        ["Last Run At", _safe_text(preventive_summary.get("last_run_at") or "-")],
    ]
    preventive_table = Table(preventive_rows, repeatRows=1, colWidths=[doc.width * 0.42, doc.width * 0.58])
    preventive_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, border),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(preventive_table)

    if preventive_runs:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Recent Preventive Runs", styles["h1"]))
        run_rows = [["Template", "Status", "Devices", "Failed Checks", "Critical"]]
        for run in preventive_runs[:6]:
            summary = dict(run.get("summary") or {})
            run_rows.append(
                [
                    _safe_text(run.get("template_name") or "Unnamed Template"),
                    _safe_text(run.get("status") or "completed"),
                    _safe_text(summary.get("devices_total") or 0),
                    _safe_text(summary.get("failed_checks_total") or 0),
                    _safe_text(summary.get("critical_devices") or 0),
                ]
            )
        run_table = Table(
            run_rows,
            repeatRows=1,
            colWidths=[doc.width * 0.40, doc.width * 0.16, doc.width * 0.14, doc.width * 0.15, doc.width * 0.15],
        )
        run_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), primary),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.35, border),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(run_table)

    elements.append(PageBreak())
    elements.append(Paragraph("Change Control Review", styles["h1"]))
    approval_rows = [
        ["Metric", "Value"],
        ["Recent Requests", _safe_text(approval_summary.get("total") or 0)],
        ["Pending", _safe_text(approval_summary.get("pending") or 0)],
        ["Approved", _safe_text(approval_summary.get("approved") or 0)],
        ["Rejected", _safe_text(approval_summary.get("rejected") or 0)],
        ["Evidence Ready", _safe_text(approval_summary.get("evidence_ready_count") or 0)],
        ["Rollback Tracked", _safe_text(approval_summary.get("rollback_tracked_count") or 0)],
    ]
    approval_table = Table(approval_rows, repeatRows=1, colWidths=[doc.width * 0.42, doc.width * 0.58])
    approval_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, border),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(approval_table)

    if approval_items:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Recent Approval Requests", styles["h1"]))
        for item in approval_items[:6]:
            status = _safe_text(item.get("execution_status") or item.get("status") or "pending")
            elements.append(
                Paragraph(
                    (
                        f"<b>#{_safe_text(item.get('id'))} {_safe_text(item.get('title') or 'Request')}</b><br/>"
                        f"<font color='{muted.hexval()}'>"
                        f"type={_safe_text(item.get('request_type') or 'approval')} | "
                        f"status={status} | requester={_safe_text(item.get('requester_name') or 'operator')}"
                        f"</font>"
                    ),
                    styles["body"],
                )
            )
            elements.append(Spacer(1, 4))

    elements.append(PageBreak())
    elements.append(Paragraph("Service Operations Review", styles["h1"]))
    if group_items:
        group_rows = [["Service Group", "Criticality", "Owner", "Members"]]
        for group in group_items:
            group_rows.append(
                [
                    _safe_text(group.get("name") or "Unnamed Group"),
                    _safe_text(group.get("criticality") or "standard"),
                    _safe_text(group.get("owner_team") or "unassigned"),
                    _safe_text(group.get("member_count") or 0),
                ]
            )
        group_table = Table(
            group_rows,
            repeatRows=1,
            colWidths=[doc.width * 0.40, doc.width * 0.18, doc.width * 0.24, doc.width * 0.18],
        )
        group_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), primary),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.35, border),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        elements.append(group_table)
        elements.append(Spacer(1, 10))

    elements.append(Paragraph("Service-Scoped Issues", styles["h1"]))
    if not issue_items:
        elements.append(Paragraph("No active service-scoped issues were captured in this review window.", styles["body"]))
    else:
        for issue in issue_items[:6]:
            service_summary = dict(issue.get("service_impact_summary") or {})
            sop_summary = dict(issue.get("sop_summary") or {})
            approval_ctx = dict(issue.get("approval_summary") or {})
            action_summary = dict(issue.get("action_summary") or {})
            knowledge_summary = dict(issue.get("knowledge_summary") or {})
            elements.append(
                Paragraph(
                    (
                        f"<b>{_safe_text(issue.get('title') or 'Issue')}</b><br/>"
                        f"<font color='{muted.hexval()}'>"
                        f"service={_safe_text(service_summary.get('primary_name') or '-')} | "
                        f"sop={_safe_text(sop_summary.get('readiness_status') or 'limited_context')} | "
                        f"actions={_safe_text(action_summary.get('total') or 0)} | "
                        f"knowledge={_safe_text(knowledge_summary.get('recommendation_count') or 0)} | "
                        f"approvals={_safe_text(approval_ctx.get('total') or 0)}"
                        f"</font>"
                    ),
                    styles["body"],
                )
            )
            elements.append(Spacer(1, 5))

    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Action Continuity Review", styles["h1"]))
    action_rows = [
        ["Metric", "Value"],
        ["Issues In Scope", _safe_text(action_summary.get("issues_in_scope") or len(action_items) or 0)],
        ["With Active Actions", _safe_text(action_summary.get("with_active_actions") or 0)],
        ["With Assigned Owners", _safe_text(action_summary.get("with_assignee") or 0)],
        ["With Knowledge Matches", _safe_text(action_summary.get("with_knowledge") or 0)],
        ["With Evidence-Linked Approvals", _safe_text(action_summary.get("with_evidence_ready") or 0)],
        ["Limited Context", _safe_text(action_summary.get("limited_context") or 0)],
    ]
    action_table = Table(action_rows, repeatRows=1, colWidths=[doc.width * 0.42, doc.width * 0.58])
    action_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, border),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(action_table)

    if action_items:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Active Continuity Items", styles["h1"]))
        for item in action_items[:6]:
            elements.append(
                Paragraph(
                    (
                        f"<b>{_safe_text(item.get('title') or 'Issue')}</b><br/>"
                        f"<font color='{muted.hexval()}'>"
                        f"service={_safe_text(item.get('primary_service') or '-')} | "
                        f"action={_safe_text(item.get('action_status') or 'open')} | "
                        f"owner={_safe_text(item.get('action_owner') or '-')} | "
                        f"knowledge={_safe_text(item.get('knowledge_matches') or 0)} | "
                        f"sop={_safe_text(item.get('sop_status') or 'limited_context')} | "
                        f"approvals={_safe_text(item.get('approval_total') or 0)}"
                        f"</font>"
                    ),
                    styles["body"],
                )
            )
            if item.get("action_note"):
                elements.append(
                    Paragraph(
                        f"<font color='{muted.hexval()}'>Latest note: {_safe_text(item.get('action_note'))}</font>",
                        styles["body"],
                    )
                )
            elements.append(Spacer(1, 4))

    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Follow-up Agenda", styles["h1"]))
    follow_up_rows = [
        ["Metric", "Value"],
        ["Needs Action Creation", _safe_text(follow_up_summary.get("needs_action") or 0)],
        ["Needs Owner Assignment", _safe_text(follow_up_summary.get("needs_owner") or 0)],
        ["Needs Knowledge Capture", _safe_text(follow_up_summary.get("needs_knowledge") or 0)],
        ["Needs Evidence Capture", _safe_text(follow_up_summary.get("needs_evidence") or 0)],
        ["Ready For Handoff", _safe_text(follow_up_summary.get("ready_for_handoff") or 0)],
    ]
    follow_up_table = Table(follow_up_rows, repeatRows=1, colWidths=[doc.width * 0.42, doc.width * 0.58])
    follow_up_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, border),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(follow_up_table)

    if follow_up_items:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Recommended Next Steps", styles["h1"]))
        for item in follow_up_items[:6]:
            elements.append(
                Paragraph(
                    (
                        f"<b>{_safe_text(item.get('title') or 'Issue')}</b><br/>"
                        f"<font color='{muted.hexval()}'>"
                        f"priority={_safe_text(item.get('priority') or 'normal')} | "
                        f"next={_safe_text(item.get('recommended_step') or 'review_and_handoff')} | "
                        f"service={_safe_text(item.get('primary_service') or '-')} | "
                        f"owner={_safe_text(item.get('action_owner') or '-')} | "
                        f"{_safe_text(item.get('step_label') or '-')}"
                        f"</font>"
                    ),
                    styles["body"],
                )
            )
            if item.get("action_note"):
                elements.append(
                    Paragraph(
                        f"<font color='{muted.hexval()}'>Latest note: {_safe_text(item.get('action_note'))}</font>",
                        styles["body"],
                    )
                )
            elements.append(Spacer(1, 4))

    elements.append(PageBreak())
    elements.append(Paragraph("Release Readiness", styles["h1"]))
    release_rows = [
        ["Metric", "Value"],
        ["Overall Status", _safe_text(release_summary.get("overall_status") or "unknown")],
        ["Accepted Gates", _safe_text(release_summary.get("accepted_gates") or 0)],
        ["Available Gates", _safe_text(release_summary.get("available_gates") or release_summary.get("total_gates") or 0)],
        ["Warnings", _safe_text(len(list(release_summary.get("warning_gates") or [])))],
        ["In Progress", _safe_text(len(list(release_summary.get("in_progress_gates") or [])))],
    ]
    release_table = Table(release_rows, repeatRows=1, colWidths=[doc.width * 0.42, doc.width * 0.58])
    release_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), primary),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, border),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, bg_soft]),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(release_table)

    if release_sections:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Release Sections", styles["h1"]))
        for section in release_sections:
            elements.append(
                Paragraph(
                    (
                        f"<b>{_safe_text(section.get('title') or section.get('id') or 'Section')}</b><br/>"
                        f"<font color='{muted.hexval()}'>"
                        f"status={_safe_text(section.get('status') or 'unknown')} | "
                        f"{_safe_text(section.get('summary') or '-')}"
                        f"</font>"
                    ),
                    styles["body"],
                )
            )
            elements.append(Spacer(1, 4))

    doc.multiBuild(elements)
    return buf.getvalue()
